import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INPUT_XLSX = ROOT / "data" / "mome_research_guide_structure.xlsx"
OUTPUT_JSON = ROOT / "data" / "research-guide.json"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def row_to_dict(row, headers):
    return {headers[i]: clean(row[i].value) for i in range(len(headers))}


def build_sections_from_card(card_row):
    items = []
    for key in ["resource_category", "search_link", "physical_location", "help_contact", "audience", "notes", "tags"]:
        value = card_row.get(key, "")
        if value:
            items.append(value)
    if not items:
        return []
    return [{"heading": "Részletek", "items": items}]


def main():
    wb = load_workbook(INPUT_XLSX, data_only=True)

    ws_nodes = wb["decision_nodes"]
    ws_cards = wb["resource_cards"]

    node_headers = [cell.value for cell in ws_nodes[1]]
    card_headers = [cell.value for cell in ws_cards[1]]

    node_rows = [row_to_dict(r, node_headers) for r in ws_nodes.iter_rows(min_row=2) if any(c.value is not None for c in r)]
    card_rows = [row_to_dict(r, card_headers) for r in ws_cards.iter_rows(min_row=2) if any(c.value is not None for c in r)]

    guide = {"startNode": "START", "nodes": {}}

    grouped = {}
    for row in node_rows:
        node_id = row.get("node_id", "")
        if not node_id:
            continue
        grouped.setdefault(node_id, []).append(row)

    for node_id, rows in grouped.items():
        first = rows[0]
        guide["nodes"][node_id] = {
            "type": "question",
            "title": first.get("question", node_id),
            "description": first.get("notes", ""),
            "options": []
        }
        for row in rows:
            option_label = row.get("option_label", "")
            next_node = row.get("next_node", "")
            if option_label and next_node:
                guide["nodes"][node_id]["options"].append({
                    "label": option_label,
                    "target": next_node
                })

    for row in card_rows:
        card_id = row.get("card_id", "")
        if not card_id:
            continue
        guide["nodes"][card_id] = {
            "type": "card",
            "title": row.get("title", card_id),
            "description": row.get("description", ""),
            "sections": build_sections_from_card(row)
        }

    OUTPUT_JSON.write_text(json.dumps(guide, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
